"""Build the BLIND human-evaluation workbook for the M6 groundedness assessment.

Step 1 of validating the M6 judge score (0.558 / 0.154) against human judgement.
Reads the judge results, the prompts sent to the judge, the phase2/phase3 answers,
and the cached truth tables; writes an .xlsx with the judge score hidden away so a
human can score the same (query, answer) pairs blind, for later judge-vs-human
agreement analysis (see score_human_eval.py).

Read-only. No LLM calls, no batch submission, no DB connection -- the cached
truth tables in m6_repair_truth_tables.json already contain the complete room
data used to build the judge's prompts, so no additional SQL pass is needed.
"""

from __future__ import annotations

import json
import random
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

REPO_ROOT = Path(__file__).resolve().parents[2]
V2_DIR = REPO_ROOT / "benchmarks" / "runs" / "phase2_vs_phase3" / "v2"
JB_DIR = V2_DIR / "judge_batches_fresh"
OUT_PATH = REPO_ROOT / "benchmarks" / "reports" / "phase2_vs_phase3" / "v2" / "m6_human_eval.xlsx"

SEED = 42
SAMPLE_TARGET = 30


def load_jsonl(path: Path) -> list[dict[str, Any]]:
    records = []
    with path.open(encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                records.append(json.loads(line))
    return records


def size_band(n: float | None) -> str:
    if n is None or n < 50:
        return "small (<50)"
    if n <= 200:
        return "medium (50-200)"
    return "large (>200)"


def format_room_line(r: dict[str, Any], relevant_attrs: list[str]) -> str:
    """Render EVERY field the JSON carries for one room. Nothing omitted.

    Earlier versions of this function rendered only flatname/roomname/zone/price
    plus a `relevant_attrs`-filtered subset of attrs, silently dropping room_id,
    house_id, neighborhood, area_m2, bed_type, deposit, deposit_value, and every
    non-"relevant" attribute. Phase3 answers routinely cite exactly those hidden
    fields (room_id, neighborhood) correctly -- the human scorer, unable to see
    them in the rendered table, mis-scored 6/14 of phase3's M6 failures as
    "fabrication"/"geographic hallucination" (see m6_step3_failure_taxonomy.md
    section 4). `relevant_attrs` is no longer used to filter anything; every
    room must be rendered complete, always.
    """
    attrs = r.get("attrs") or {}
    fields = [
        f"room_id={r.get('room_id')}",
        f"house_id={r.get('house_id')}",
        f"flatname={r.get('flatname')}",
        f"roomname={r.get('roomname')}",
        f"city={r.get('city')}",
        f"zone={r.get('zone')}",
        f"neighborhood={r.get('neighborhood')}",
        f"price_eur={r.get('price_eur')}",
        f"spring_price={r.get('spring_price')}",
        f"summer_price={r.get('summer_price')}",
        f"fixed_price={r.get('fixed_price')}",
        f"area_m2={r.get('area_m2')}",
        f"bed_type={r.get('bed_type')}",
        f"deposit={r.get('deposit')}",
        f"deposit_value={r.get('deposit_value')}",
    ]
    for k in sorted(attrs.keys()):
        fields.append(f"{k}={attrs[k]}")
    return " | ".join(fields)


def render_truth_table(tt: dict[str, Any] | None) -> tuple[str, int, str]:
    """Human-readable rendering of a truth table. Returns (summary, n_truth_rooms, band)."""
    if tt is None:
        return "No truth table available for this query_id.", 0, size_band(0)

    kind = tt.get("kind")
    total = tt.get("total_matches")
    note = (tt.get("note") or "").strip()
    lines = [f"City: {tt.get('city', '?')}"]

    if kind == "non_filterable" or total is None:
        lines.append(
            "This query is NOT answerable via a filtered SQL lookup "
            "(no truth table of rooms applies)."
        )
        if note:
            lines.append(f"Note: {note}")
        return "\n".join(lines), 0, size_band(None)

    if kind == "zone_enum":
        zones = tt.get("zones") or []
        lines.append(f"True answer: {total} zones (ground truth is a ZONE NAME LIST, not a room list).")
        for z in zones:
            lines.append(f"  - {z['zone'].strip()}: {z['n_rooms']} rooms")
        if note:
            lines.append(f"Note: {note}")
        return "\n".join(lines), total, size_band(total)

    rooms = tt.get("rooms") or tt.get("rooms_full") or []
    relevant_attrs = tt.get("relevant_attrs") or []
    lines.append(
        f"True total_matches: {total} -- COMPLETE: every matching row and every "
        "stored field is shown. Nothing has been omitted."
    )
    if total == 0:
        lines.append("Zero rooms match these filters.")
    for r in rooms:
        lines.append(f"  - {format_room_line(r, relevant_attrs)}")
    if note:
        lines.append(f"Note: {note}")
    return "\n".join(lines), total, size_band(total)


def extract_score_and_rationale(text: str) -> tuple[float | None, str, bool]:
    """Parse a judge response; handle max_tokens-truncated JSON gracefully."""
    m = re.search(r"\{.*\}", text, re.DOTALL)
    if m:
        try:
            obj = json.loads(m.group(0))
            return float(obj["score"]), obj.get("rationale", ""), False
        except Exception:
            pass
    score_m = re.search(r'"score"\s*:\s*([0-9.]+)', text)
    score = float(score_m.group(1)) if score_m else None
    rat_m = re.search(r'"rationale"\s*:\s*"(.*)$', text, re.DOTALL)
    rationale = (rat_m.group(1) if rat_m else text) + "  [TRUNCATED at max_tokens]"
    return score, rationale, True


def extract_system_prompt(batch_records: list[dict[str, Any]]) -> str:
    return batch_records[0]["params"]["system"]


def extract_not_captured_note(batch_records: list[dict[str, Any]]) -> str:
    content = batch_records[0]["params"]["messages"][0]["content"]
    idx = content.find("Fields NOT captured")
    end = content.find("\n\n**Automated fact-check")
    if idx != -1 and end != -1:
        return content[idx:end].strip()
    return ""


def build_rows() -> tuple[list[dict[str, Any]], str, str]:
    id_map_records = load_jsonl(JB_DIR / "id_mapping_M6_repaired.jsonl")
    id_map = {r["custom_id"]: r for r in id_map_records}

    results = {}
    for r in load_jsonl(JB_DIR / "results" / "results_M6_repaired.jsonl"):
        score, rationale, truncated = extract_score_and_rationale(r["text"])
        results[r["custom_id"]] = {"score": score, "rationale": rationale, "truncated": truncated}

    batch_records = load_jsonl(JB_DIR / "batch_M6_repaired.jsonl")
    system_prompt = extract_system_prompt(batch_records)
    not_captured_note = extract_not_captured_note(batch_records)

    truth_tables = json.loads((JB_DIR / "m6_repair_truth_tables.json").read_text(encoding="utf-8"))

    phase_answers: dict[tuple[str, str], dict[str, Any]] = {}
    for fname in ("phase2_eval_v2_fresh.jsonl", "phase3_eval_v2_fresh.jsonl"):
        for rec in load_jsonl(V2_DIR / fname):
            phase_answers[(rec["system"], rec["id"])] = rec

    rows = []
    for cid, m in id_map.items():
        system = m["system"]
        query_id = m["query_id"]
        res = results.get(cid, {})
        ans_rec = phase_answers.get((system, query_id))
        if ans_rec is None:
            raise RuntimeError(f"No phase answer found for {system}/{query_id}")
        tt = truth_tables.get(query_id)
        summary, n_rooms, band = render_truth_table(tt)

        rows.append(
            {
                "query_id": query_id,
                "category": m["category"],
                "language": m["language"],
                "query": ans_rec.get("query", ""),
                "answer": ans_rec.get("final_message", ""),
                "truth_table_summary": summary,
                "n_truth_rooms": n_rooms,
                "truth_table_size_band": band,
                "system": system,
                "judge_score": res.get("score"),
                "judge_rationale": res.get("rationale"),
                "judge_truncated": res.get("truncated", False),
                "m6_det": m.get("m6_det"),
            }
        )

    return rows, system_prompt, not_captured_note


def select_sample(rows: list[dict[str, Any]], rng: random.Random) -> tuple[set[tuple[str, str]], list[str]]:
    """Return the set of (query_id, system) keys in the sample, plus an audit log."""
    audit: list[str] = []

    high_risk = [
        r for r in rows if r["judge_score"] == 0.0 and r["m6_det"] is not None and r["m6_det"] >= 0.8
    ]
    high_risk_keys = {(r["query_id"], r["system"]) for r in high_risk}
    for r in high_risk:
        audit.append(
            f"HIGH-RISK  {r['query_id']:32s} {r['system']:7s} "
            f"judge=0.0 vs deterministic={r['m6_det']:.2f}"
        )

    remaining = [r for r in rows if (r["query_id"], r["system"]) not in high_risk_keys]
    quota_needed = SAMPLE_TARGET - len(high_risk)

    current_system_count = {
        "phase2": sum(1 for r in high_risk if r["system"] == "phase2"),
        "phase3": sum(1 for r in high_risk if r["system"] == "phase3"),
    }
    target_per_system = SAMPLE_TARGET // 2
    system_quota = {s: max(0, target_per_system - current_system_count[s]) for s in ("phase2", "phase3")}
    diff = quota_needed - sum(system_quota.values())
    if diff != 0:
        system_quota["phase3"] += diff  # deterministic tie-break

    selected_extra: list[dict[str, Any]] = []
    for system, quota in system_quota.items():
        pool = [r for r in remaining if r["system"] == system]
        groups: dict[tuple[str, str], list[dict[str, Any]]] = {}
        for r in pool:
            groups.setdefault((r["category"], r["truth_table_size_band"]), []).append(r)
        for key in groups:
            rng.shuffle(groups[key])
        group_keys = sorted(groups.keys())
        rng.shuffle(group_keys)

        picked: list[dict[str, Any]] = []
        idx = 0
        while len(picked) < quota and any(groups[k] for k in group_keys):
            key = group_keys[idx % len(group_keys)]
            if groups[key]:
                r = groups[key].pop()
                picked.append(r)
                audit.append(
                    f"STRAT-FILL {r['query_id']:32s} {r['system']:7s} "
                    f"category={r['category']} band={r['truth_table_size_band']}"
                )
            idx += 1
        selected_extra.extend(picked)

    sample_keys = high_risk_keys | {(r["query_id"], r["system"]) for r in selected_extra}
    return sample_keys, audit


HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
WRAP_TOP = Alignment(wrap_text=True, vertical="top")


def style_header(ws, headers: list[str], widths: list[int]) -> None:
    for col_idx, (name, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP_TOP
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_workbook(
    rows: list[dict[str, Any]],
    sample_keys: set[tuple[str, str]],
    system_prompt: str,
    not_captured_note: str,
) -> None:
    wb = Workbook()

    # ---- Sheet 1: eval ----
    ws = wb.active
    ws.title = "eval"
    headers = [
        "row_id", "query_id", "category", "language", "query", "answer",
        "truth_table_summary", "n_truth_rooms", "truth_table_size_band",
        "in_sample_30", "human_score", "human_note",
    ]
    widths = [8, 30, 22, 10, 45, 70, 70, 14, 18, 14, 12, 40]
    style_header(ws, headers, widths)

    for row_idx, r in enumerate(rows, start=2):
        in_sample = (r["query_id"], r["system"]) in sample_keys
        values = [
            r["row_id"], r["query_id"], r["category"], r["language"], r["query"], r["answer"],
            r["truth_table_summary"], r["n_truth_rooms"], r["truth_table_size_band"],
            in_sample, None, None,
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = WRAP_TOP
        max_chars = max(len(str(v)) for v in values if v is not None)
        ws.row_dimensions[row_idx].height = max(30, min(300, max_chars // 4 * 5))

    dv = DataValidation(type="list", formula1='"0,0.5,1"', allow_blank=True, showErrorMessage=True)
    dv.error = "human_score must be 0, 0.5, or 1"
    ws.add_data_validation(dv)
    dv.add(f"K2:K{len(rows) + 1}")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:L{len(rows) + 1}"

    # ---- Sheet 2: rubric ----
    ws2 = wb.create_sheet("rubric")
    ws2.column_dimensions["A"].width = 110
    r = 1
    ws2.cell(row=r, column=1, value="Judge system prompt (verbatim)").font = Font(bold=True, size=12)
    r += 1
    cell = ws2.cell(row=r, column=1, value=system_prompt)
    cell.alignment = WRAP_TOP
    ws2.row_dimensions[r].height = 600
    r += 2
    ws2.cell(row=r, column=1, value="Attributes the truth table does NOT carry (verbatim, per-query note)").font = Font(
        bold=True, size=12
    )
    r += 1
    cell = ws2.cell(row=r, column=1, value=not_captured_note)
    cell.alignment = WRAP_TOP
    ws2.row_dimensions[r].height = 90

    # ---- Sheet 3: _key (hidden) ----
    ws3 = wb.create_sheet("_key")
    key_headers = ["row_id", "query_id", "system", "judge_score", "judge_rationale", "judge_truncated"]
    for col_idx, name in enumerate(key_headers, start=1):
        ws3.cell(row=1, column=col_idx, value=name).font = Font(bold=True)
    for row_idx, r in enumerate(rows, start=2):
        values = [r["row_id"], r["query_id"], r["system"], r["judge_score"], r["judge_rationale"], r["judge_truncated"]]
        for col_idx, value in enumerate(values, start=1):
            ws3.cell(row=row_idx, column=col_idx, value=value)
    ws3.sheet_state = "hidden"

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)


def main() -> int:
    rows, system_prompt, not_captured_note = build_rows()

    rng = random.Random(SEED)
    sample_keys, audit = select_sample(rows, rng)

    rng.shuffle(rows)
    for i, r in enumerate(rows, start=1):
        r["row_id"] = i

    write_workbook(rows, sample_keys, system_prompt, not_captured_note)

    band_counts: dict[str, int] = {}
    for r in rows:
        band_counts[r["truth_table_size_band"]] = band_counts.get(r["truth_table_size_band"], 0) + 1

    sample_rows = [r for r in rows if (r["query_id"], r["system"]) in sample_keys]
    sample_band_counts: dict[str, int] = {}
    sample_cat_counts: dict[str, int] = {}
    sample_sys_counts: dict[str, int] = {}
    for r in sample_rows:
        sample_band_counts[r["truth_table_size_band"]] = sample_band_counts.get(r["truth_table_size_band"], 0) + 1
        sample_cat_counts[r["category"]] = sample_cat_counts.get(r["category"], 0) + 1
        sample_sys_counts[r["system"]] = sample_sys_counts.get(r["system"], 0) + 1

    print(f"Workbook written to: {OUT_PATH}")
    print(f"RNG seed: {SEED}")
    print(f"Rows written: {len(rows)}")
    print(f"Full-set band distribution: {band_counts}")
    print(f"Sample size: {len(sample_rows)}")
    print(f"Sample system balance: {sample_sys_counts}")
    print(f"Sample category distribution: {sample_cat_counts}")
    print(f"Sample band distribution: {sample_band_counts}")
    print()
    print("Sample audit log (row selection reasons):")
    for line in audit:
        print(f"  {line}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
