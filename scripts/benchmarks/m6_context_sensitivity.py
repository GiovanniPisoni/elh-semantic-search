"""Step 2 of the M6 assessment: does the judge degrade as the truth table grows?

Read-only analysis of existing files -- no API calls, no DB writes. Joins the
judge's actual measured input-token count (usage.in from the batch results,
i.e. what really happened when the batch ran) against the human scores
collected in m6_human_eval.xlsx, and checks whether judge-vs-human divergence
tracks truth-table size, and whether that is separable from query difficulty
(human score also dropping with size).

Writes benchmarks/reports/phase2_vs_phase3/v2/m6_step2_context_sensitivity.md
"""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from scipy import stats

REPO_ROOT = Path(__file__).resolve().parents[2]
V2_DIR = REPO_ROOT / "benchmarks" / "runs" / "phase2_vs_phase3" / "v2"
JB_DIR = V2_DIR / "judge_batches_fresh"
WORKBOOK_PATH = REPO_ROOT / "benchmarks" / "reports" / "phase2_vs_phase3" / "v2" / "m6_human_eval.xlsx"
OUT_PATH = REPO_ROOT / "benchmarks" / "reports" / "phase2_vs_phase3" / "v2" / "m6_step2_context_sensitivity.md"

LEVELS = (0.0, 0.5, 1.0)
AGREE_THRESHOLD = 0.25


def load_jsonl(path: Path) -> list[dict[str, Any]]:
    out = []
    with path.open(encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                out.append(json.loads(line))
    return out


def extract_score(text: str) -> tuple[float | None, bool]:
    """Parse judge score; handle max_tokens-truncated JSON (closing brace missing)."""
    m = re.search(r"\{.*\}", text, re.DOTALL)
    if m:
        try:
            obj = json.loads(m.group(0))
            return float(obj["score"]), False
        except Exception:
            pass
    score_m = re.search(r'"score"\s*:\s*([0-9.]+)', text)
    score = float(score_m.group(1)) if score_m else None
    return score, True


def size_band(n: float | None) -> str:
    if n is None or n < 50:
        return "small (<50)"
    if n <= 200:
        return "medium (50-200)"
    return "large (>200)"


def cohens_kappa(pairs: list[tuple[float, float]]) -> float | None:
    n = len(pairs)
    if n == 0:
        return None
    idx = {lv: i for i, lv in enumerate(LEVELS)}
    confusion = [[0] * len(LEVELS) for _ in LEVELS]
    for judge, human in pairs:
        confusion[idx[judge]][idx[human]] += 1
    observed = sum(confusion[i][i] for i in range(len(LEVELS))) / n
    row_t = [sum(confusion[i]) for i in range(len(LEVELS))]
    col_t = [sum(confusion[i][j] for i in range(len(LEVELS))) for j in range(len(LEVELS))]
    expected = sum(row_t[i] * col_t[i] for i in range(len(LEVELS))) / (n * n)
    if expected == 1.0:
        return 1.0
    return (observed - expected) / (1 - expected)


def build_records() -> list[dict[str, Any]]:
    id_map = {r["custom_id"]: r for r in load_jsonl(JB_DIR / "id_mapping_M6_repaired.jsonl")}

    results = {}
    for r in load_jsonl(JB_DIR / "results" / "results_M6_repaired.jsonl"):
        score, truncated = extract_score(r["text"])
        results[r["custom_id"]] = {
            "score": score,
            "truncated": truncated,
            "prompt_tokens": r["usage"]["in"],
        }

    truth_tables = json.loads((JB_DIR / "m6_repair_truth_tables.json").read_text(encoding="utf-8"))

    wb = load_workbook(WORKBOOK_PATH, data_only=True)
    ws_eval = wb["eval"]
    ws_key = wb["_key"]
    eval_headers = [c.value for c in ws_eval[1]]
    eval_idx = {name: i for i, name in enumerate(eval_headers)}
    key_headers = [c.value for c in ws_key[1]]
    key_idx = {name: i for i, name in enumerate(key_headers)}

    human_by_row_id = {}
    for row in ws_eval.iter_rows(min_row=2, values_only=True):
        rid = row[eval_idx["row_id"]]
        if rid is None:
            continue
        human_by_row_id[rid] = row[eval_idx["human_score"]]

    human_by_key = {}
    for row in ws_key.iter_rows(min_row=2, values_only=True):
        rid = row[key_idx["row_id"]]
        if rid is None:
            continue
        query_id = row[key_idx["query_id"]]
        system = row[key_idx["system"]]
        raw = human_by_row_id.get(rid)
        human_by_key[(query_id, system)] = float(raw) if raw is not None else None

    records = []
    for cid, m in id_map.items():
        system = m["system"]
        query_id = m["query_id"]
        res = results[cid]
        tt = truth_tables.get(query_id) or {}
        n_truth_rooms = tt.get("total_matches")
        if n_truth_rooms is None:
            n_truth_rooms = 0
        human_score = human_by_key.get((query_id, system))
        judge_score = res["score"]

        records.append(
            {
                "query_id": query_id,
                "system": system,
                "category": m["category"],
                "n_truth_rooms": n_truth_rooms,
                "prompt_tokens": res["prompt_tokens"],
                "judge_score": judge_score,
                "human_score": human_score,
                "delta": judge_score - human_score,
                "abs_delta": abs(judge_score - human_score),
                "agree": abs(judge_score - human_score) < AGREE_THRESHOLD,
                "truncated": res["truncated"],
                "size_band": size_band(n_truth_rooms),
            }
        )
    return records


def token_quartile_bins(records: list[dict[str, Any]]) -> dict[str, tuple[float, float]]:
    tokens = sorted(r["prompt_tokens"] for r in records)
    q1, q2, q3 = (
        stats.scoreatpercentile(tokens, 25),
        stats.scoreatpercentile(tokens, 50),
        stats.scoreatpercentile(tokens, 75),
    )
    return {"Q1": (0, q1), "Q2": (q1, q2), "Q3": (q2, q3), "Q4": (q3, float("inf"))}, (q1, q2, q3)


def assign_quartile(tok: float, boundaries: tuple[float, float, float]) -> str:
    q1, q2, q3 = boundaries
    if tok <= q1:
        return "Q1"
    if tok <= q2:
        return "Q2"
    if tok <= q3:
        return "Q3"
    return "Q4"


def bin_stats(recs: list[dict[str, Any]]) -> dict[str, Any]:
    n = len(recs)
    if n == 0:
        return {"n": 0}
    mean_judge = sum(r["judge_score"] for r in recs) / n
    mean_human = sum(r["human_score"] for r in recs) / n
    mean_delta = sum(r["delta"] for r in recs) / n
    agree_pct = 100.0 * sum(1 for r in recs if r["agree"]) / n
    pairs = [(r["judge_score"], r["human_score"]) for r in recs]
    kappa = cohens_kappa(pairs)
    return {
        "n": n,
        "mean_judge": mean_judge,
        "mean_human": mean_human,
        "mean_delta": mean_delta,
        "agree_pct": agree_pct,
        "kappa": kappa,
    }


def fmt(v, nd=3):
    if v is None:
        return "n/a"
    if isinstance(v, float):
        return f"{v:.{nd}f}"
    return str(v)


def main() -> int:
    records = build_records()
    n_total = len(records)

    # ---- size bins ----
    band_order = ["small (<50)", "medium (50-200)", "large (>200)"]
    by_band = {b: [r for r in records if r["size_band"] == b] for b in band_order}

    # ---- token quartiles ----
    _, boundaries = token_quartile_bins(records)
    for r in records:
        r["token_quartile"] = assign_quartile(r["prompt_tokens"], boundaries)
    quartile_order = ["Q1", "Q2", "Q3", "Q4"]
    by_quartile = {q: [r for r in records if r["token_quartile"] == q] for q in quartile_order}

    # ---- correlations ----
    def corr(xs, ys):
        rho, p = stats.spearmanr(xs, ys)
        return rho, p, len(xs)

    n_rooms = [r["n_truth_rooms"] for r in records]
    tokens = [r["prompt_tokens"] for r in records]
    judge_scores = [r["judge_score"] for r in records]
    human_scores = [r["human_score"] for r in records]
    deltas = [r["delta"] for r in records]
    abs_deltas = [r["abs_delta"] for r in records]

    corr_rooms_judge = corr(n_rooms, judge_scores)
    corr_rooms_human = corr(n_rooms, human_scores)
    corr_rooms_delta = corr(n_rooms, deltas)
    corr_rooms_absdelta = corr(n_rooms, abs_deltas)

    corr_tok_judge = corr(tokens, judge_scores)
    corr_tok_human = corr(tokens, human_scores)
    corr_tok_delta = corr(tokens, deltas)
    corr_tok_absdelta = corr(tokens, abs_deltas)

    # ---- truncation distribution ----
    truncated_recs = [r for r in records if r["truncated"]]
    trunc_by_band = {b: sum(1 for r in by_band[b] if r["truncated"]) for b in band_order}
    trunc_by_quartile = {q: sum(1 for r in by_quartile[q] if r["truncated"]) for q in quartile_order}

    # ---- 0.5 usage per bin ----
    half_by_band = {
        b: (
            sum(1 for r in by_band[b] if r["judge_score"] == 0.5),
            sum(1 for r in by_band[b] if r["human_score"] == 0.5),
            len(by_band[b]),
        )
        for b in band_order
    }

    # ============ build markdown ============
    lines: list[str] = []
    lines.append("# M6 Step 2 — Judge context-sensitivity analysis")
    lines.append("")
    lines.append(
        "Read-only analysis joining judge scores, human scores (from `m6_human_eval.xlsx`), "
        "and the truth-table size / actual prompt token count of all 52 M6-repaired judge calls. "
        "No API calls were made: prompt token counts are the **actual measured `usage.in`** values "
        "recorded in `results_M6_repaired.jsonl` when the batch was run, not an estimate."
    )
    lines.append("")

    # --- 1. per-record table ---
    lines.append("## 1. Per-record data")
    lines.append("")
    lines.append(
        "| query_id | system | n_truth_rooms | prompt_tokens | judge | human | delta | agree | truncated |"
    )
    lines.append("|---|---|---:|---:|---:|---:|---:|:---:|:---:|")
    for r in sorted(records, key=lambda x: x["n_truth_rooms"]):
        lines.append(
            f"| {r['query_id']} | {r['system']} | {r['n_truth_rooms']} | {r['prompt_tokens']} | "
            f"{fmt(r['judge_score'],1)} | {fmt(r['human_score'],1)} | {fmt(r['delta'],2)} | "
            f"{'Y' if r['agree'] else 'N'} | {'Y' if r['truncated'] else ''} |"
        )
    lines.append("")

    # --- 2. bin by row count ---
    lines.append("## 2. Binned by truth-table row count (n_truth_rooms)")
    lines.append("")
    lines.append("| bin | n | mean judge | mean human | mean delta (judge-human) | exact agreement % | kappa |")
    lines.append("|---|---:|---:|---:|---:|---:|---:|")
    for b in band_order:
        s = bin_stats(by_band[b])
        lines.append(
            f"| {b} | {s['n']} | {fmt(s.get('mean_judge'))} | {fmt(s.get('mean_human'))} | "
            f"{fmt(s.get('mean_delta'))} | {fmt(s.get('agree_pct'),1)} | {fmt(s.get('kappa'))} |"
        )
    lines.append("")

    # --- 3. bin by token quartile ---
    lines.append("## 3. Binned by prompt-token quartile (actual usage.in)")
    lines.append("")
    lines.append(
        f"Quartile boundaries (tokens): Q1 <= {boundaries[0]:.0f} < Q2 <= {boundaries[1]:.0f} "
        f"< Q3 <= {boundaries[2]:.0f} < Q4"
    )
    lines.append("")
    lines.append("| bin | n | token range | mean judge | mean human | mean delta | exact agreement % | kappa |")
    lines.append("|---|---:|---|---:|---:|---:|---:|---:|")
    edges = [0] + list(boundaries) + [max(tokens)]
    for i, q in enumerate(quartile_order):
        s = bin_stats(by_quartile[q])
        lines.append(
            f"| {q} | {s['n']} | {edges[i]:.0f}-{edges[i+1]:.0f} | {fmt(s.get('mean_judge'))} | "
            f"{fmt(s.get('mean_human'))} | {fmt(s.get('mean_delta'))} | {fmt(s.get('agree_pct'),1)} | "
            f"{fmt(s.get('kappa'))} |"
        )
    lines.append("")

    # --- 4. correlations ---
    lines.append("## 4. Correlations (Spearman, n=52)")
    lines.append("")
    lines.append("| x | y | rho | p | significant (p<0.05)? |")
    lines.append("|---|---|---:|---:|:---:|")

    def corr_row(name_x, name_y, c):
        rho, p, n = c
        sig = "yes" if p < 0.05 else "no"
        lines.append(f"| {name_x} | {name_y} | {fmt(rho)} | {fmt(p,4)} | {sig} |")

    corr_row("n_truth_rooms", "judge_score", corr_rooms_judge)
    corr_row("n_truth_rooms", "human_score", corr_rooms_human)
    corr_row("n_truth_rooms", "delta (judge-human)", corr_rooms_delta)
    corr_row("n_truth_rooms", "|delta|", corr_rooms_absdelta)
    corr_row("prompt_tokens", "judge_score", corr_tok_judge)
    corr_row("prompt_tokens", "human_score", corr_tok_human)
    corr_row("prompt_tokens", "delta (judge-human)", corr_tok_delta)
    corr_row("prompt_tokens", "|delta|", corr_tok_absdelta)
    lines.append("")
    lines.append(
        f"n=52 for the whole-sample correlations. Per-bin subgroup stats above (section 2/3) have "
        f"n as low as {min(bin_stats(by_band[b])['n'] for b in band_order)} "
        "(large-table bin) — treat those subgroup numbers as descriptive, not statistically decisive."
    )
    lines.append("")

    # --- 5. confound check ---
    lines.append("## 5. Confound check — does the human score also drop with size?")
    lines.append("")
    lines.append("| predictor | outcome | rho | p |")
    lines.append("|---|---|---:|---:|")
    lines.append(
        f"| n_truth_rooms | human_score | {fmt(corr_rooms_human[0])} | {fmt(corr_rooms_human[1],4)} |"
    )
    lines.append(
        f"| n_truth_rooms | judge_score | {fmt(corr_rooms_judge[0])} | {fmt(corr_rooms_judge[1],4)} |"
    )
    lines.append(
        f"| prompt_tokens | human_score | {fmt(corr_tok_human[0])} | {fmt(corr_tok_human[1],4)} |"
    )
    lines.append(
        f"| prompt_tokens | judge_score | {fmt(corr_tok_judge[0])} | {fmt(corr_tok_judge[1],4)} |"
    )
    lines.append("")

    # --- 6. truncation check ---
    lines.append("## 6. Truncation check (10/52 judge responses hit max_tokens)")
    lines.append("")
    lines.append("| bin | truncated | total | % truncated |")
    lines.append("|---|---:|---:|---:|")
    for b in band_order:
        t, tot = trunc_by_band[b], len(by_band[b])
        lines.append(f"| {b} | {t} | {tot} | {fmt(100*t/tot,1) if tot else 'n/a'} |")
    lines.append("")
    lines.append("| token quartile | truncated | total | % truncated |")
    lines.append("|---|---:|---:|---:|")
    for q in quartile_order:
        t, tot = trunc_by_quartile[q], len(by_quartile[q])
        lines.append(f"| {q} | {t} | {tot} | {fmt(100*t/tot,1) if tot else 'n/a'} |")
    lines.append("")
    lines.append(f"Overall truncation rate: {len(truncated_recs)}/{n_total} = {100*len(truncated_recs)/n_total:.1f}%.")
    lines.append("")

    # --- 7. 0.5 usage per bin ---
    lines.append("## 7. Judge vs human use of the 0.5 (hedge) score, per size bin")
    lines.append("")
    lines.append("| bin | judge=0.5 | human=0.5 | n |")
    lines.append("|---|---:|---:|---:|")
    for b in band_order:
        j_half, h_half, tot = half_by_band[b]
        lines.append(f"| {b} | {j_half} | {h_half} | {tot} |")
    total_j_half = sum(1 for r in records if r["judge_score"] == 0.5)
    total_h_half = sum(1 for r in records if r["human_score"] == 0.5)
    lines.append(f"| **all** | **{total_j_half}** | **{total_h_half}** | **{n_total}** |")
    lines.append("")

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text("\n".join(lines), encoding="utf-8")

    # ---- console summary ----
    print(f"Report written to: {OUT_PATH}")
    print(f"n={n_total}")
    print("Band sizes:", {b: len(by_band[b]) for b in band_order})
    print("Quartile boundaries (tokens):", [round(x) for x in boundaries])
    print()
    print("Correlations (n_truth_rooms):")
    print("  vs judge_score:", corr_rooms_judge)
    print("  vs human_score:", corr_rooms_human)
    print("  vs delta:      ", corr_rooms_delta)
    print("  vs |delta|:    ", corr_rooms_absdelta)
    print("Correlations (prompt_tokens):")
    print("  vs judge_score:", corr_tok_judge)
    print("  vs human_score:", corr_tok_human)
    print("  vs delta:      ", corr_tok_delta)
    print("  vs |delta|:    ", corr_tok_absdelta)
    print()
    print("Truncation by band:", trunc_by_band, "| by quartile:", trunc_by_quartile)
    print("0.5 usage by band (judge, human, n):", half_by_band)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
