"""Assemble the joined per-record dataset for the M6 Step-3 failure taxonomy.

Read-only. Joins answers, judge scores/rationales, human scores/notes, and a
readable truth-table summary for all 52 M6-repaired records into one JSON file
for manual (human-in-the-loop / LLM analyst) classification. Does not classify
anything itself -- classification requires reading the answer text against the
truth table, which is not mechanical.
"""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[2]
V2_DIR = REPO_ROOT / "benchmarks" / "runs" / "phase2_vs_phase3" / "v2"
JB_DIR = V2_DIR / "judge_batches_fresh"
WORKBOOK_PATH = REPO_ROOT / "benchmarks" / "reports" / "phase2_vs_phase3" / "v2" / "m6_human_eval.xlsx"
OUT_PATH = REPO_ROOT / "benchmarks" / "reports" / "phase2_vs_phase3" / "v2" / "_m6_failure_data.json"

MAX_TRUTH_ROWS_SHOWN = 40


def load_jsonl(path: Path) -> list[dict[str, Any]]:
    out = []
    with path.open(encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                out.append(json.loads(line))
    return out


def extract_score_and_rationale(text: str) -> tuple[float | None, str, bool]:
    m = re.search(r"\{.*\}", text, re.DOTALL)
    if m:
        try:
            obj = json.loads(m.group(0))
            return float(obj["score"]), obj.get("rationale", ""), False
        except Exception:
            pass
    score_m = re.search(r'"score"\s*:\s*([0-9.]+)', text)
    score = float(score_m.group(1)) if score_m else None
    rat_m = re.search(r'"rationale"\s*:\s*"(.*)$', text, re.DOTALL)
    rationale = (rat_m.group(1) if rat_m else text) + "  [TRUNCATED at max_tokens]"
    return score, rationale, True


def format_price(r: dict[str, Any]) -> str:
    if r.get("fixed_price"):
        return f"EUR{r['price_eur']:.0f} (fixed)"
    return f"EUR{r['price_eur']:.0f} aut / EUR{r['spring_price']:.0f} spr / EUR{r['summer_price']:.0f} sum"


def format_room_line(r: dict[str, Any], relevant_attrs: list[str]) -> str:
    attrs = r.get("attrs") or {}
    key_attrs = [f"{a}={attrs[a]}" for a in relevant_attrs if a in attrs]
    attrs_str = f" | {', '.join(key_attrs)}" if key_attrs else ""
    return f"{r['flatname']} ({r['roomname']}) | {r['zone']} | {format_price(r)}{attrs_str}"


def render_truth_table(tt: dict[str, Any] | None) -> str:
    if tt is None:
        return "No truth table available."
    kind = tt.get("kind")
    total = tt.get("total_matches")
    note = (tt.get("note") or "").strip()
    lines = [f"City: {tt.get('city', '?')}"]
    if kind == "non_filterable" or total is None:
        lines.append("NOT answerable via filtered SQL lookup (no truth table applies).")
        if note:
            lines.append(f"Note: {note}")
        return "\n".join(lines)
    if kind == "zone_enum":
        zones = tt.get("zones") or []
        lines.append(f"True answer: {total} zones (ground truth is a ZONE NAME LIST, not a room list).")
        for z in zones:
            lines.append(f"  - {z['zone'].strip()}: {z['n_rooms']} rooms")
        if note:
            lines.append(f"Note: {note}")
        return "\n".join(lines)
    rooms = tt.get("rooms") or tt.get("rooms_full") or []
    relevant_attrs = tt.get("relevant_attrs") or []
    lines.append(f"True total_matches: {total}")
    if total == 0:
        lines.append("Zero rooms match these filters.")
    shown = rooms[:MAX_TRUTH_ROWS_SHOWN]
    for r in shown:
        lines.append(f"  - {format_room_line(r, relevant_attrs)}")
    if len(rooms) > MAX_TRUTH_ROWS_SHOWN:
        lines.append(f"  ... and {len(rooms) - MAX_TRUTH_ROWS_SHOWN} more rooms (all match the filters)")
    if note:
        lines.append(f"Note: {note}")
    return "\n".join(lines)


def main() -> int:
    id_map = {r["custom_id"]: r for r in load_jsonl(JB_DIR / "id_mapping_M6_repaired.jsonl")}

    results = {}
    for r in load_jsonl(JB_DIR / "results" / "results_M6_repaired.jsonl"):
        score, rationale, truncated = extract_score_and_rationale(r["text"])
        results[r["custom_id"]] = {"score": score, "rationale": rationale, "truncated": truncated}

    truth_tables = json.loads((JB_DIR / "m6_repair_truth_tables.json").read_text(encoding="utf-8"))

    phase_answers: dict[tuple[str, str], dict[str, Any]] = {}
    for fname in ("phase2_eval_v2_fresh.jsonl", "phase3_eval_v2_fresh.jsonl"):
        for rec in load_jsonl(V2_DIR / fname):
            phase_answers[(rec["system"], rec["id"])] = rec

    wb = load_workbook(WORKBOOK_PATH, data_only=True)
    ws_eval = wb["eval"]
    ws_key = wb["_key"]
    eval_headers = [c.value for c in ws_eval[1]]
    eval_idx = {name: i for i, name in enumerate(eval_headers)}
    key_headers = [c.value for c in ws_key[1]]
    key_idx = {name: i for i, name in enumerate(key_headers)}

    human_by_row_id = {}
    for row in ws_eval.iter_rows(min_row=2, values_only=True):
        rid = row[eval_idx["row_id"]]
        if rid is None:
            continue
        human_by_row_id[rid] = {
            "human_score": row[eval_idx["human_score"]],
            "human_note": row[eval_idx["human_note"]],
        }

    human_by_key: dict[tuple[str, str], dict[str, Any]] = {}
    for row in ws_key.iter_rows(min_row=2, values_only=True):
        rid = row[key_idx["row_id"]]
        if rid is None:
            continue
        query_id = row[key_idx["query_id"]]
        system = row[key_idx["system"]]
        h = human_by_row_id.get(rid, {})
        raw = h.get("human_score")
        human_by_key[(query_id, system)] = {
            "human_score": float(raw) if raw is not None else None,
            "human_note": h.get("human_note"),
        }

    records = []
    for cid, m in id_map.items():
        system = m["system"]
        query_id = m["query_id"]
        res = results[cid]
        ans_rec = phase_answers[(system, query_id)]
        tt = truth_tables.get(query_id)
        human = human_by_key.get((query_id, system), {})

        records.append(
            {
                "query_id": query_id,
                "system": system,
                "category": m["category"],
                "query": ans_rec.get("query", ""),
                "answer": ans_rec.get("final_message", ""),
                "human_score": human.get("human_score"),
                "human_note": human.get("human_note"),
                "judge_score": res["score"],
                "judge_rationale": res["rationale"],
                "judge_truncated": res["truncated"],
                "n_truth_rooms": (tt or {}).get("total_matches"),
                "truth_table_summary": render_truth_table(tt),
            }
        )

    records.sort(key=lambda r: (r["query_id"], r["system"]))

    OUT_PATH.write_text(json.dumps(records, indent=2, ensure_ascii=False), encoding="utf-8")
    n_fail = sum(1 for r in records if r["human_score"] is not None and r["human_score"] < 1.0)
    print(f"Wrote {len(records)} records to {OUT_PATH}")
    print(f"Failures (human_score < 1.0): {n_fail}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
