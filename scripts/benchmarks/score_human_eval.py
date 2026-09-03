"""Score judge-vs-human agreement on the M6 blind human-evaluation workbook.

Reads benchmarks/reports/phase2_vs_phase3/v2/m6_human_eval.xlsx once human_score
has been filled in on the "eval" sheet, joins it against the hidden "_key" sheet
by row_id, and prints agreement statistics.
"""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_PATH = REPO_ROOT / "benchmarks" / "reports" / "phase2_vs_phase3" / "v2" / "m6_human_eval.xlsx"

LEVELS = (0.0, 0.5, 1.0)


def load_rows(path: Path) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    ws_eval = wb["eval"]
    ws_key = wb["_key"]

    eval_headers = [c.value for c in ws_eval[1]]
    eval_idx = {name: i for i, name in enumerate(eval_headers)}

    key_headers = [c.value for c in ws_key[1]]
    key_idx = {name: i for i, name in enumerate(key_headers)}

    key_by_row_id = {}
    for row in ws_key.iter_rows(min_row=2, values_only=True):
        if row[key_idx["row_id"]] is None:
            continue
        key_by_row_id[row[key_idx["row_id"]]] = {
            "system": row[key_idx["system"]],
            "judge_score": row[key_idx["judge_score"]],
            "judge_rationale": row[key_idx["judge_rationale"]],
            "judge_truncated": row[key_idx["judge_truncated"]],
        }

    rows = []
    for row in ws_eval.iter_rows(min_row=2, values_only=True):
        row_id = row[eval_idx["row_id"]]
        if row_id is None:
            continue
        k = key_by_row_id.get(row_id, {})
        rows.append(
            {
                "row_id": row_id,
                "query_id": row[eval_idx["query_id"]],
                "category": row[eval_idx["category"]],
                "in_sample_30": row[eval_idx["in_sample_30"]],
                "human_score": row[eval_idx["human_score"]],
                "human_note": row[eval_idx["human_note"]],
                "system": k.get("system"),
                "judge_score": k.get("judge_score"),
                "judge_rationale": k.get("judge_rationale"),
                "judge_truncated": k.get("judge_truncated"),
            }
        )
    return rows


def nearest_level(x: float) -> float:
    return min(LEVELS, key=lambda lv: abs(lv - x))


def cohens_kappa(pairs: list[tuple[float, float]]) -> float:
    n = len(pairs)
    if n == 0:
        return float("nan")
    idx = {lv: i for i, lv in enumerate(LEVELS)}
    confusion = [[0] * len(LEVELS) for _ in LEVELS]
    for judge, human in pairs:
        confusion[idx[judge]][idx[human]] += 1

    observed_agreement = sum(confusion[i][i] for i in range(len(LEVELS))) / n

    row_totals = [sum(confusion[i]) for i in range(len(LEVELS))]
    col_totals = [sum(confusion[i][j] for i in range(len(LEVELS))) for j in range(len(LEVELS))]
    expected_agreement = sum(row_totals[i] * col_totals[i] for i in range(len(LEVELS))) / (n * n)

    if expected_agreement == 1.0:
        return 1.0
    return (observed_agreement - expected_agreement) / (1 - expected_agreement)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--path", type=Path, default=DEFAULT_PATH, help="Path to m6_human_eval.xlsx")
    args = parser.parse_args()

    rows = load_rows(args.path)
    scored = [
        r
        for r in rows
        if r["human_score"] is not None and r["judge_score"] is not None
    ]

    print(f"Workbook: {args.path}")
    print(f"Total rows: {len(rows)}")
    print(f"Rows scored by human so far: {len(scored)}")
    print()

    if not scored:
        print("No rows scored yet -- fill in human_score on the 'eval' sheet and re-run.")
        return 0

    pairs = []
    for r in scored:
        judge = nearest_level(float(r["judge_score"]))
        human = nearest_level(float(r["human_score"]))
        pairs.append((judge, human))

    n = len(pairs)
    exact_matches = sum(1 for j, h in pairs if j == h)
    exact_agreement_pct = 100.0 * exact_matches / n
    kappa = cohens_kappa(pairs)

    mean_judge = sum(j for j, _ in pairs) / n
    mean_human = sum(h for _, h in pairs) / n

    print(f"Judge-vs-human exact agreement: {exact_agreement_pct:.1f}% ({exact_matches}/{n})")
    print(f"Cohen's kappa: {kappa:.3f}")
    print(f"Mean judge score (scored subset): {mean_judge:.3f}")
    print(f"Mean human score (scored subset): {mean_human:.3f}")
    print()

    print("Confusion matrix (rows=judge, cols=human):")
    header = "judge\\human".ljust(12) + "".join(f"{lv:>8}" for lv in LEVELS)
    print(header)
    idx = {lv: i for i, lv in enumerate(LEVELS)}
    confusion = [[0] * len(LEVELS) for _ in LEVELS]
    for judge, human in pairs:
        confusion[idx[judge]][idx[human]] += 1
    for i, lv in enumerate(LEVELS):
        print(f"{lv:<12}" + "".join(f"{confusion[i][j]:>8}" for j in range(len(LEVELS))))
    print()

    disagreements = [
        r
        for r in scored
        if abs(float(r["judge_score"]) - float(r["human_score"])) >= 0.5
    ]
    disagreements.sort(key=lambda r: -abs(float(r["judge_score"]) - float(r["human_score"])))

    print(f"Rows disagreeing by >= 0.5 ({len(disagreements)}):")
    for r in disagreements:
        note = r["human_note"] or ""
        trunc = " [judge response truncated]" if r["judge_truncated"] else ""
        print(
            f"  row_id={r['row_id']:<4} query_id={r['query_id']:<32} system={r['system']:<7} "
            f"judge={r['judge_score']} human={r['human_score']}{trunc}"
        )
        if note:
            print(f"    human_note: {note}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
