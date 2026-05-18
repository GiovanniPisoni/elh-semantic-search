"""Generate an Excel template for human evaluation of an agent benchmark run.

Reads a JSONL file produced by :mod:`scripts.benchmarks.run_agent_benchmark`,
strips the technical fields, and writes an Excel file with one row per query.
"""

from __future__ import annotations

import argparse
import json
import logging
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from elh_rag.logging_setup import setup_logging

logger = logging.getLogger(__name__)


DEFAULT_RUNS_DIR = Path("benchmarks/runs")
DEFAULT_EVAL_DIR = Path("benchmarks/human_eval")


COLUMN_SPEC: list[tuple[str, int]] = [
    ("id", 18),
    ("category", 14),
    ("language", 9),
    ("query", 50),
    ("agent_response", 80),
    ("tools_used", 30),
    ("correctness_1_10", 14),
    ("completeness_1_10", 14),
    ("comments", 50),
]


def latest_run_file(runs_dir: Path) -> Path:
    """Return the most recently modified JSONL file in runs_dir."""
    if not runs_dir.exists():
        raise FileNotFoundError(f"Runs directory does not exist: {runs_dir}")
    candidates = sorted(runs_dir.glob("agent_benchmark_*.jsonl"), key=lambda p: p.stat().st_mtime)
    if not candidates:
        raise FileNotFoundError(f"No agent_benchmark_*.jsonl files in {runs_dir}")
    return candidates[-1]


def load_jsonl(path: Path) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                records.append(json.loads(line))
    return records


def write_workbook(records: list[dict[str, Any]], output_path: Path) -> None:
    """Build and save the Excel workbook."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Human Evaluation"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    instruction_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    text_align = Alignment(wrap_text=True, vertical="top")

    for col_idx, (name, width) in enumerate(COLUMN_SPEC, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = text_align
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    instructions = {
        "correctness_1_10": "1 = fully wrong, 10 = fully accurate",
        "completeness_1_10": "1 = missing key info, 10 = nothing to add",
        "comments": "Free text: errors, omissions, language quality",
    }
    for col_idx, (name, _) in enumerate(COLUMN_SPEC, start=1):
        if name in instructions:
            cell = ws.cell(row=2, column=col_idx, value=instructions[name])
            cell.fill = instruction_fill
            cell.alignment = text_align
            cell.font = Font(italic=True, size=9)

    for row_idx, record in enumerate(records, start=3):
        if record.get("status") != "success":
            response = f"[FAILED: {record.get('error', 'unknown')}]"
            tools = ""
        else:
            response = record.get("final_message", "")
            tools = ", ".join(record.get("tools_used", []))
        values = [
            record.get("id", ""),
            record.get("category", ""),
            record.get("language", ""),
            record.get("query", ""),
            response,
            tools,
            "",  # correctness_1_10
            "",  # completeness_1_10
            "",  # comments
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = text_align

        # Approximate row height proportional to longest cell content.
        max_chars = max(len(str(v)) for v in values)
        ws.row_dimensions[row_idx].height = max(45, min(300, max_chars // 4 * 5))

    ws.freeze_panes = "A3"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--input",
        type=Path,
        default=None,
        help="JSONL run file. If omitted, the most recent file in benchmarks/runs/ is used.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_EVAL_DIR,
        help=f"Directory for the generated Excel (default: {DEFAULT_EVAL_DIR}).",
    )
    return parser.parse_args()


def main() -> int:
    setup_logging()
    args = parse_args()

    input_path = args.input or latest_run_file(DEFAULT_RUNS_DIR)
    print(f"Reading: {input_path}")
    records = load_jsonl(input_path)

    output_name = input_path.stem.replace("agent_benchmark_", "human_eval_") + ".xlsx"
    output_path = args.output_dir / output_name
    write_workbook(records, output_path)

    print(f"Excel template written to: {output_path}")
    print(f"Queries: {len(records)}")
    print("Send this file to the domain expert for evaluation.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
